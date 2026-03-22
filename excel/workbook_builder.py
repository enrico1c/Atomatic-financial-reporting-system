"""
Excel workbook builder using openpyxl.

Sheet layout:
  1. RAW_PRICES       — raw daily prices from Yahoo Finance
  2. RAW_MACRO        — raw macro from FRED + ECB
  3. CLEAN_PRICES     — cleaned daily prices + returns + volatility
  4. CLEAN_MACRO      — cleaned monthly macro indicators
  5. FINANCIAL_STMTS  — per-ticker income stmt / balance sheet / cash flow
  6. KPI_SUMMARY      — calculated KPIs (growth, volatility, spreads)
  7. MONTHLY_REPORT   — monthly aggregated view of prices + macro
  8. INSIGHTS         — rule-based text insights
  9. WORLDBANK        — annual World Bank macro data
"""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from config.settings import OUTPUT_DIR, LATEST_EXCEL_FILENAME, REPORT_EXCEL_FILENAME
from utils.logger import get_logger

log = get_logger("excel.workbook_builder")

# ── Color palette ──────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
LIGHT_GREY = "F2F2F2"
GREEN      = "375623"
RED_BG     = "FFC7CE"
GREEN_BG   = "C6EFCE"
YELLOW_BG  = "FFEB9C"


class WorkbookBuilder:
    def __init__(self, clean_data: dict[str, pd.DataFrame], insights: list[str]):
        self.data     = clean_data
        self.insights = insights
        self.wb       = Workbook()
        self.wb.remove(self.wb.active)  # remove default sheet

    # ── Public ─────────────────────────────────────────────────────────────────
    def build(self) -> Path:
        log.info("Building Excel workbook")
        self._sheet_raw_prices()
        self._sheet_raw_macro()
        self._sheet_clean_prices()
        self._sheet_clean_macro()
        self._sheet_financial_stmts()
        self._sheet_kpi_summary()
        self._sheet_monthly_report()
        self._sheet_insights()
        self._sheet_worldbank()
        path = self._save()
        return path

    # ── Sheet builders ─────────────────────────────────────────────────────────
    def _sheet_raw_prices(self):
        df = self.data.get("prices_daily")
        if df is None or df.empty:
            return
        # Only show OHLCV columns (not returns/vols) for the RAW sheet
        ohlcv_cols = [c for c in df.columns if any(
            c.endswith(s) for s in ["_open", "_high", "_low", "_close", "_volume"]
        )]
        self._write_df(df[ohlcv_cols].tail(252), "RAW_PRICES", header_color=MID_BLUE)

    def _sheet_raw_macro(self):
        df = self.data.get("macro_monthly")
        if df is None or df.empty:
            return
        # Only base indicator columns, no growth suffixes
        base_cols = [c for c in df.columns if not c.endswith(("_mom_1m", "_yoy"))]
        self._write_df(df[base_cols].tail(60), "RAW_MACRO", header_color=MID_BLUE)

    def _sheet_clean_prices(self):
        df = self.data.get("prices_daily")
        if df is None or df.empty:
            return
        self._write_df(df.tail(504), "CLEAN_PRICES", header_color=DARK_BLUE, fmt_numbers=True)

    def _sheet_clean_macro(self):
        df = self.data.get("macro_monthly")
        if df is None or df.empty:
            return
        self._write_df(df.tail(120), "CLEAN_MACRO", header_color=DARK_BLUE, fmt_numbers=True)

    def _sheet_financial_stmts(self):
        ws = self.wb.create_sheet("FINANCIAL_STMTS")
        self._style_tab(ws, DARK_BLUE)
        row_offset = 1

        from config.settings import STATEMENT_TICKERS
        for ticker in STATEMENT_TICKERS:
            for period in ["annual", "quarterly"]:
                key = f"financials_{ticker}_{period}"
                df = self.data.get(key)
                if df is None or df.empty:
                    continue
                # Section header
                ws.cell(row=row_offset, column=1).value = f"── {ticker} | {period.upper()} ──"
                self._style_header_cell(ws.cell(row=row_offset, column=1), DARK_BLUE)
                row_offset += 1
                self._append_df_to_ws(ws, df.tail(8), row_offset)
                row_offset += len(df.tail(8)) + 3

        self._autofit(ws)

    def _sheet_kpi_summary(self):
        df = self.data.get("kpi_summary")
        if df is None or df.empty:
            return
        self._write_df(df, "KPI_SUMMARY", header_color=DARK_BLUE, fmt_numbers=True)

    def _sheet_monthly_report(self):
        df = self.data.get("monthly_report")
        if df is None or df.empty:
            return
        self._write_df(df.tail(24), "MONTHLY_REPORT", header_color=DARK_BLUE, fmt_numbers=True)

    def _sheet_insights(self):
        ws = self.wb.create_sheet("INSIGHTS")
        self._style_tab(ws, DARK_BLUE)

        # Title
        title = ws.cell(row=1, column=1, value="Financial Insights Summary")
        title.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
        title.fill = PatternFill("solid", fgColor=DARK_BLUE)
        ws.merge_cells("A1:D1")
        ws.row_dimensions[1].height = 30

        ts = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ts.font = Font(italic=True, color="555555")
        ws.merge_cells("A2:D2")

        ws.cell(row=3, column=1)  # spacer

        for i, insight in enumerate(self.insights, start=4):
            cell = ws.cell(row=i, column=1, value=insight)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(wrap_text=True)
            # Color-code by severity keyword
            if any(w in insight.lower() for w in ["elevated", "high", "tighten", "recession", "inver"]):
                cell.fill = PatternFill("solid", fgColor=RED_BG)
            elif any(w in insight.lower() for w in ["strong", "increas", "growth", "positive"]):
                cell.fill = PatternFill("solid", fgColor=GREEN_BG)
            else:
                cell.fill = PatternFill("solid", fgColor=YELLOW_BG)
            ws.merge_cells(f"A{i}:D{i}")
            ws.row_dimensions[i].height = 28

        ws.column_dimensions["A"].width = 90

    def _sheet_worldbank(self):
        df = self.data.get("macro_worldbank")
        if df is None or df.empty:
            return
        self._write_df(df, "WORLDBANK", header_color=MID_BLUE)

    # ── Generic helpers ────────────────────────────────────────────────────────
    def _write_df(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        header_color: str = MID_BLUE,
        fmt_numbers: bool = False,
    ):
        ws = self.wb.create_sheet(sheet_name)
        self._style_tab(ws, header_color)
        self._append_df_to_ws(ws, df, start_row=1, header_color=header_color, fmt_numbers=fmt_numbers)
        self._autofit(ws)

    def _append_df_to_ws(
        self,
        ws,
        df: pd.DataFrame,
        start_row: int = 1,
        header_color: str = MID_BLUE,
        fmt_numbers: bool = False,
    ):
        df = df.reset_index()
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    self._style_header_cell(cell, header_color)
                else:
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")
                    if r_idx % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
                    if fmt_numbers and isinstance(value, float):
                        cell.number_format = "#,##0.00"

    def _style_header_cell(self, cell, color: str = MID_BLUE):
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _style_tab(self, ws, color: str):
        ws.sheet_properties.tabColor = color

    def _autofit(self, ws, max_width: int = 30, min_width: int = 8):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

    # ── Save ───────────────────────────────────────────────────────────────────
    def _save(self) -> Path:
        dated_name = REPORT_EXCEL_FILENAME.format(date=datetime.now().strftime("%Y%m%d"))
        dated_path = OUTPUT_DIR / dated_name
        latest_path = OUTPUT_DIR / LATEST_EXCEL_FILENAME

        self.wb.save(dated_path)
        self.wb.save(latest_path)  # always overwrite latest

        log.info(f"Workbook saved: {dated_path}")
        log.info(f"Latest copy:    {latest_path}")
        return latest_path
